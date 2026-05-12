from django.shortcuts import render

# Create your views here.
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from transactions.models import Transaction
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import io

@login_required
def reports_index(request):
    return render(request, 'reports/index.html')

@login_required
def export_excel(request):
    user = request.user
    transactions = Transaction.objects.filter(user=user).order_by('-date')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'รายการธุรกรรม'

    # Header style
    header_fill = PatternFill(start_color='0D1117', end_color='0D1117', fill_type='solid')
    header_font = Font(bold=True, color='00D4FF', size=11)
    header_align = Alignment(horizontal='center', vertical='center')

    headers = ['วันที่', 'ประเภท', 'หมวดหมู่', 'กระเป๋าเงิน', 'จำนวนเงิน', 'หมายเหตุ', 'แท็ก']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 20

    # Data rows
    for row, t in enumerate(transactions, 2):
        ws.cell(row=row, column=1, value=str(t.date))
        ws.cell(row=row, column=2, value='รายรับ' if t.transaction_type == 'income' else 'รายจ่าย')
        ws.cell(row=row, column=3, value=t.category.name if t.category else '-')
        ws.cell(row=row, column=4, value=t.wallet.wallet_name if t.wallet else '-')
        ws.cell(row=row, column=5, value=float(t.amount))
        ws.cell(row=row, column=6, value=t.note or '-')
        ws.cell(row=row, column=7, value=t.tags or '-')

        # สีแถว
        if t.transaction_type == 'income':
            fill = PatternFill(start_color='0D2B1A', end_color='0D2B1A', fill_type='solid')
            font = Font(color='00FF88')
        else:
            fill = PatternFill(start_color='2B0D0D', end_color='2B0D0D', fill_type='solid')
            font = Font(color='FF4757')

        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = fill
        ws.cell(row=row, column=5).font = font

    # Summary sheet
    ws2 = wb.create_sheet('สรุป')
    today = timezone.now().date()

    total_income = sum(float(t.amount) for t in transactions if t.transaction_type == 'income')
    total_expense = sum(float(t.amount) for t in transactions if t.transaction_type == 'expense')

    summary_data = [
        ['รายงานสรุปการเงิน', ''],
        ['ผู้ใช้', user.username],
        ['วันที่ออกรายงาน', str(today)],
        ['', ''],
        ['รายรับรวม', total_income],
        ['รายจ่ายรวม', total_expense],
        ['คงเหลือสุทธิ', total_income - total_expense],
        ['จำนวนรายการ', len(transactions)],
    ]

    for row_data in summary_data:
        ws2.append(row_data)

    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 20

    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="finance_report_{today}.xlsx"'
    wb.save(response)
    return response

@login_required
def export_pdf(request):
    user = request.user
    transactions = Transaction.objects.filter(user=user).order_by('-date')[:50]
    today = timezone.now().date()

    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    # Title
    p.setFillColorRGB(0, 0.83, 1)
    p.setFont('Helvetica-Bold', 18)
    p.drawString(50, height - 60, 'SmartFinance Report')

    p.setFillColorRGB(0, 0, 0)
    p.setFont('Helvetica', 11)
    p.drawString(50, height - 85, f'User: {user.username}  |  Date: {today}')

    p.setStrokeColorRGB(0.8, 0.8, 0.8)
    p.line(50, height - 100, width - 50, height - 100)

    # Summary
    total_income = sum(float(t.amount) for t in Transaction.objects.filter(user=user, transaction_type='income'))
    total_expense = sum(float(t.amount) for t in Transaction.objects.filter(user=user, transaction_type='expense'))

    p.setFillColorRGB(0, 0, 0)
    p.setFont('Helvetica-Bold', 12)
    p.drawString(50, height - 130, 'Summary')

    p.setFont('Helvetica', 11)
    p.setFillColorRGB(0, 0.6, 0.2)
    p.drawString(50, height - 155, f'Total Income:   {total_income:,.2f} THB')
    p.setFillColorRGB(0.8, 0.1, 0.1)
    p.drawString(50, height - 175, f'Total Expense:  {total_expense:,.2f} THB')
    p.setFillColorRGB(0, 0, 0)
    p.drawString(50, height - 195, f'Net Balance:    {total_income - total_expense:,.2f} THB')

    # Table header
    p.setFillColorRGB(0.07, 0.07, 0.1)
    p.rect(50, height - 230, width - 100, 20, fill=1, stroke=0)

    p.setFillColorRGB(0, 0.83, 1)
    p.setFont('Helvetica-Bold', 9)
    p.drawString(55, height - 224, 'Date')
    p.drawString(130, height - 224, 'Type')
    p.drawString(210, height - 224, 'Category')
    p.drawString(330, height - 224, 'Amount (THB)')
    p.drawString(430, height - 224, 'Note')

    # Table rows
    y = height - 250
    for t in transactions:
        if y < 60:
            p.showPage()
            y = height - 60

        p.setFont('Helvetica', 8)

        # Type
        if t.transaction_type == 'income':
            p.setFillColorRGB(0, 0.6, 0.2)
            type_text = 'Income'
        else:
            p.setFillColorRGB(0.8, 0.1, 0.1)
            type_text = 'Expense'

        p.drawString(55, y, str(t.date))
        p.drawString(130, y, type_text)

        # Category — แปลงเป็น ASCII
        p.setFillColorRGB(0, 0, 0)
        if t.category:
            try:
                cat_name = t.category.name.encode('ascii').decode('ascii')
            except (UnicodeEncodeError, UnicodeDecodeError):
                cat_name = f'Cat-{t.category.pk}'
        else:
            cat_name = '-'

        # Note — แปลงเป็น ASCII
        if t.note:
            try:
                note_text = t.note.encode('ascii').decode('ascii')
            except (UnicodeEncodeError, UnicodeDecodeError):
                note_text = '-'
        else:
            note_text = '-'

        p.drawString(210, y, cat_name[:18])
        p.drawString(330, y, f'{float(t.amount):,.2f}')
        p.drawString(430, y, note_text[:18])

        p.setStrokeColorRGB(0.8, 0.8, 0.8)
        p.line(50, y - 4, width - 50, y - 4)
        y -= 18

    p.save()
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="finance_report_{today}.pdf"'
    return response